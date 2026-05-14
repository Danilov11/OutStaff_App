#!/usr/bin/env python3
"""
Импорт данных из Excel в Supabase PostgreSQL
Файл: Сырье_панель_Чайхана-2.xlsx

Установить зависимости:
  pip install openpyxl psycopg2-binary

Запуск:
  python3 supabase/import_data.py
"""

import psycopg2
import openpyxl
from datetime import datetime
import sys
import os

# ── Настройки подключения ────────────────────────────────────
CONN_STR = (
    "postgresql://postgres.gjhcvsmaatwrzortwdms:"
    "pahniq-ritqa4-Fiqkag@aws-0-eu-west-1.pooler.supabase.com:5432/postgres"
)

EXCEL_PATH = os.path.join(os.path.dirname(__file__), '..', 'Сырье_панель_Чайхана-2.xlsx')
# Если файл не рядом с проектом, укажи путь явно:
# EXCEL_PATH = '/Users/admin/Downloads/Сырье_панель_Чайхана-2.xlsx'


# ── Вспомогательные функции ──────────────────────────────────

def fmt_inn(val):
    """ИНН хранится как большое float (774342493800.0) → строка '774342493800'"""
    if val is None or str(val).strip() in ('', 'None'):
        return None
    try:
        return str(int(float(str(val))))
    except (ValueError, OverflowError):
        return str(val).strip() or None

def fmt_date(val):
    """datetime → 'DD.MM.YYYY', строка возвращается как есть"""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime('%d.%m.%Y')
    s = str(val).strip()
    return s if s else None

def fmt_phone(val):
    """Телефон может прийти как float (79252580102.0) → строка"""
    if val is None:
        return None
    s = str(val).strip()
    if s.endswith('.0'):
        s = s[:-2]
    return s if s else None

def fmt_num(val):
    """Число или None"""
    if val is None:
        return None
    try:
        return float(val)
    except (TypeError, ValueError):
        return None

def clean(val):
    """Строка или None"""
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


# ── Загрузка листов ──────────────────────────────────────────

def load_payments(ws):
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        # Год, Период, Сотрудник, Телефон, Сумма из реестра, Удержание,
        # Итог с удержанием, Комиссия платформы, ИТОГ к выплате, Статус, Комментарий, ИНН
        year = row[0]
        if not year:
            continue
        period   = clean(row[1])
        employee = clean(row[2])
        if not employee:
            continue
        rows.append({
            'year':                 int(year),
            'period':               period or '',
            'employee':             employee,
            'phone':                fmt_phone(row[3]),
            'amount_registry':      fmt_num(row[4]),
            'deduction':            fmt_num(row[5]),
            'total_with_deduction': fmt_num(row[6]),
            'platform_fee':         fmt_num(row[7]),
            'amount':               fmt_num(row[8]) or 0,
            'status':               clean(row[9]) or '',
            'comment':              clean(row[10]),
            'inn':                  fmt_inn(row[11]),
        })
    return rows


def load_documents(ws):
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        # Статус, Проект, Город, Должность, Ресторан, Комментарий,
        # Дата проверки в РКЛ, Отпуск, ИНН, серия патента, Номер патента,
        # Серия бланка, Номер бланка, Дата выдачи паспорта, Дата рождения,
        # Паспортные данные, ФИО, Телефон, Гражданство, Ссылка на документы,
        # Проблемы, Регистрация дата окончания, Патент дата выдачи,
        # Договор дата, Ссылка на договор, Уволен (дата), ИП Договоры
        employee = clean(row[16])
        if not employee:
            continue
        rows.append({
            'status':               clean(row[0]),
            'project':              clean(row[1]),
            'city':                 clean(row[2]),
            'position':             clean(row[3]),
            'restaurant':           clean(row[4]),
            'comment':              clean(row[5]),
            'rcl_check_date':       fmt_date(row[6]),
            'vacation':             clean(row[7]),
            'inn':                  fmt_inn(row[8]),
            'patent_series':        clean(row[9]),
            'patent_number':        clean(row[10]),
            'patent_blank_series':  clean(row[11]),
            'patent_blank_number':  clean(row[12]),
            'passport_issue_date':  fmt_date(row[13]),
            'birth_date':           fmt_date(row[14]),
            'passport_data':        clean(row[15]),
            'employee':             employee,
            'phone':                fmt_phone(row[17]),
            'citizenship':          clean(row[18]),
            'documents_link':       clean(row[19]),
            'problems':             clean(row[20]),
            'registration_end_date': fmt_date(row[21]),
            'patent_issue_date':    fmt_date(row[22]),
            'contract_date':        fmt_date(row[23]),
            'contract_link':        clean(row[24]),
            'dismissed_date':       fmt_date(row[25]),
            'ip_contracts':         clean(row[26]),
        })
    return rows


def load_rounds(ws):
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        # id, createdAt, restaurant, director, employeeName, employeeInn,
        # paperReason, plannedVisitDate, status, adminDeadline,
        # adminComment, contractDeliveryDate, updatedAt
        rid = clean(row[0])
        if not rid:
            continue
        rows.append({
            'id':                     rid,
            'created_at':             row[1] if isinstance(row[1], datetime) else None,
            'restaurant':             clean(row[2]),
            'director':               clean(row[3]),
            'employee_name':          clean(row[4]),
            'employee_inn':           fmt_inn(row[5]),
            'paper_reason':           clean(row[6]),
            'planned_visit_date':     fmt_date(row[7]),
            'status':                 clean(row[8]),
            'admin_deadline':         fmt_date(row[9]),
            'admin_comment':          clean(row[10]),
            'contract_delivery_date': fmt_date(row[11]),
            'updated_at':             row[12] if isinstance(row[12], datetime) else None,
        })
    return rows


def load_accounts(ws):
    """Лист Счета: Год, Период, Месяц, ФОТ, Выручка, Оплачено, Разница, Статус, Комментарий"""
    rows = []
    transactions = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        year   = row[0]
        period = clean(row[1])
        if not year or not period:
            continue
        rows.append({
            'year':         int(year),
            'period':       period,
            'month':        clean(row[2]),
            'payroll_fund': fmt_num(row[3]),
            'revenue':      fmt_num(row[4]),
            'amount_paid':  fmt_num(row[5]) or 0,
            'status':       clean(row[7]),
            'comment':      clean(row[8]),
        })
        # Транзакции: Лицо (col 12), Дата (col 13), Сумма (col 14)
        person = clean(row[12])
        date   = fmt_date(row[13])
        amount = fmt_num(row[14])
        if person and date and amount:
            transactions.append({
                'date':   date,
                'person': person,
                'amount': amount,
            })
    return rows, transactions


# ── Вставка данных ───────────────────────────────────────────

def insert_payments(cur, rows):
    cur.executemany("""
        INSERT INTO payments
          (year, period, employee, phone, inn,
           amount_registry, deduction, total_with_deduction, platform_fee,
           amount, status, comment)
        VALUES
          (%(year)s, %(period)s, %(employee)s, %(phone)s, %(inn)s,
           %(amount_registry)s, %(deduction)s, %(total_with_deduction)s, %(platform_fee)s,
           %(amount)s, %(status)s, %(comment)s)
    """, rows)
    print(f'  Выплаты: вставлено {len(rows)} записей')


def insert_documents(cur, rows):
    cur.executemany("""
        INSERT INTO documents
          (status, project, city, position, restaurant, comment,
           rcl_check_date, vacation, inn,
           patent_series, patent_number, patent_blank_series, patent_blank_number,
           passport_issue_date, birth_date, passport_data,
           employee, phone, citizenship, documents_link, problems,
           registration_end_date, patent_issue_date, contract_date, contract_link,
           dismissed_date, ip_contracts)
        VALUES
          (%(status)s, %(project)s, %(city)s, %(position)s, %(restaurant)s, %(comment)s,
           %(rcl_check_date)s, %(vacation)s, %(inn)s,
           %(patent_series)s, %(patent_number)s, %(patent_blank_series)s, %(patent_blank_number)s,
           %(passport_issue_date)s, %(birth_date)s, %(passport_data)s,
           %(employee)s, %(phone)s, %(citizenship)s, %(documents_link)s, %(problems)s,
           %(registration_end_date)s, %(patent_issue_date)s, %(contract_date)s, %(contract_link)s,
           %(dismissed_date)s, %(ip_contracts)s)
    """, rows)
    print(f'  Документы: вставлено {len(rows)} записей')


def insert_rounds(cur, rows):
    cur.executemany("""
        INSERT INTO rounds
          (id, created_at, restaurant, director, employee_name, employee_inn,
           paper_reason, planned_visit_date, status,
           admin_deadline, admin_comment, contract_delivery_date, updated_at)
        VALUES
          (%(id)s, %(created_at)s, %(restaurant)s, %(director)s, %(employee_name)s, %(employee_inn)s,
           %(paper_reason)s, %(planned_visit_date)s, %(status)s,
           %(admin_deadline)s, %(admin_comment)s, %(contract_delivery_date)s, %(updated_at)s)
        ON CONFLICT (id) DO NOTHING
    """, rows)
    print(f'  Объезды: вставлено {len(rows)} записей')


def insert_accounts(cur, rows):
    cur.executemany("""
        INSERT INTO account_payments
          (year, period, month, payroll_fund, revenue, amount_paid, status, comment)
        VALUES
          (%(year)s, %(period)s, %(month)s, %(payroll_fund)s, %(revenue)s,
           %(amount_paid)s, %(status)s, %(comment)s)
    """, rows)
    print(f'  Счета: вставлено {len(rows)} записей')


def insert_transactions(cur, rows):
    # Убираем дубликаты
    seen = set()
    unique = []
    for r in rows:
        key = (r['date'], r['person'], r['amount'])
        if key not in seen:
            seen.add(key)
            unique.append(r)
    cur.executemany("""
        INSERT INTO transactions (date, person, amount)
        VALUES (%(date)s, %(person)s, %(amount)s)
    """, unique)
    print(f'  Транзакции: вставлено {len(unique)} записей')


# ── Главная функция ──────────────────────────────────────────

def main():
    excel_path = EXCEL_PATH
    if not os.path.exists(excel_path):
        # Попробовать в текущей папке
        alt = '/Users/admin/Downloads/Сырье_панель_Чайхана-2.xlsx'
        if os.path.exists(alt):
            excel_path = alt
        else:
            print(f'Файл не найден: {excel_path}')
            sys.exit(1)

    print(f'Читаем файл: {excel_path}')
    wb = openpyxl.load_workbook(excel_path, data_only=True)

    print('Подготовка данных...')
    payments     = load_payments(wb['Выплаты'])
    documents    = load_documents(wb['Документы'])
    rounds       = load_rounds(wb['Объезды'])
    accounts, transactions = load_accounts(wb['Счета'])

    print(f'  Выплат:      {len(payments)}')
    print(f'  Документов:  {len(documents)}')
    print(f'  Объездов:    {len(rounds)}')
    print(f'  Счетов:      {len(accounts)}')
    print(f'  Транзакций:  {len(transactions)}')

    print('\nПодключение к Supabase...')
    conn = psycopg2.connect(CONN_STR)
    conn.autocommit = False
    cur = conn.cursor()

    try:
        print('Очистка таблиц...')
        cur.execute('TRUNCATE payments, documents, rounds, account_payments, transactions RESTART IDENTITY CASCADE')

        print('Вставка данных...')
        insert_payments(cur, payments)
        insert_documents(cur, documents)
        insert_rounds(cur, rounds)
        insert_accounts(cur, accounts)
        insert_transactions(cur, transactions)

        conn.commit()
        print('\n✅ Импорт завершён успешно!')

    except Exception as e:
        conn.rollback()
        print(f'\n❌ Ошибка: {e}')
        raise
    finally:
        cur.close()
        conn.close()


if __name__ == '__main__':
    main()
